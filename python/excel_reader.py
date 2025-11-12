"""
Excel Reader class to simplify reading and extracting data from Excel files
"""
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional


class ExcelReader:
    """Excel reader class to simplify reading and extracting data from Excel files"""

    def __init__(self, file_path: str):
        """
        Initialize ExcelReader with file path
        
        Args:
            file_path: Path to the Excel file
        """
        self.file_path = file_path
        self.workbook = None

    def load(self) -> None:
        """Load the Excel file"""
        self.workbook = load_workbook(filename=self.file_path, data_only=True)

    def get_worksheet(self, sheet_name: str):
        """
        Get a worksheet by name
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            Worksheet object or None if not found
        """
        if self.workbook is None:
            raise ValueError("Workbook not loaded. Call load() first.")
        
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        return None

    def extract_columns(
        self,
        sheet_name: str,
        columns: List[str],
        start_row: int = 2
    ) -> List[Dict[str, Any]]:
        """
        Extract data from specific columns in a worksheet
        
        Args:
            sheet_name: Name of the sheet
            columns: Column letters to extract (e.g., ['A', 'Q'])
            start_row: Starting row (default: 2, skip header)
            
        Returns:
            Array of row data with specified columns
        """
        worksheet = self.get_worksheet(sheet_name)
        if worksheet is None:
            raise ValueError(f'Worksheet "{sheet_name}" not found')

        data = []

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=start_row), start=start_row):
            row_data = {}
            
            for col in columns:
                # Convert column letter to index (A=0, B=1, etc.)
                col_idx = self._column_letter_to_index(col)
                if col_idx < len(row):
                    row_data[col] = row[col_idx].value
                else:
                    row_data[col] = None

            # Only add row if at least one column has data
            if any(val is not None for val in row_data.values()):
                row_data['_rowNumber'] = row_idx
                data.append(row_data)

        return data

    def get_sheet_names(self) -> List[str]:
        """
        Get all sheet names in the workbook
        
        Returns:
            List of sheet names
        """
        if self.workbook is None:
            raise ValueError("Workbook not loaded. Call load() first.")
        return self.workbook.sheetnames

    def get_row_count(self, sheet_name: str) -> int:
        """
        Get row count for a specific sheet
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            Number of rows
        """
        worksheet = self.get_worksheet(sheet_name)
        if worksheet is None:
            raise ValueError(f'Worksheet "{sheet_name}" not found')
        return worksheet.max_row

    @staticmethod
    def _column_letter_to_index(column: str) -> int:
        """
        Convert column letter to zero-based index
        
        Args:
            column: Column letter (e.g., 'A', 'B', 'AA')
            
        Returns:
            Zero-based column index
        """
        column = column.upper()
        index = 0
        for char in column:
            index = index * 26 + (ord(char) - ord('A') + 1)
        return index - 1
